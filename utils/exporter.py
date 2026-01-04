import io
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# ===============================
# MAIN EXCEL EXPORT (ENHANCED)
# ===============================
def export_excel(df, job_role="Resume Screening"):
    """
    Enhanced Excel export with:
    - Professional formatting
    - Color-coded scores
    - Auto-adjusted column widths
    - Headers with styling
    - Metadata sheet
    """
    buffer = io.BytesIO()
    
    # Create Excel with multiple sheets
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        # Main results sheet
        df[
            [
                "Candidate",
                "Matching Percentage",
                "Phone",
                "Email",
                "Matched Skills",
                "Missing Skills"
            ]
        ].to_excel(writer, sheet_name="Results", index=False)
        
        # Summary statistics sheet
        create_summary_sheet(writer, df, job_role)
        
        # Get workbook to apply formatting
        workbook = writer.book
        worksheet = workbook["Results"]
        
        # Apply professional styling
        apply_excel_formatting(worksheet, df)
    
    buffer.seek(0)
    return buffer

# ===============================
# EXCEL FORMATTING 
# ===============================
def apply_excel_formatting(worksheet, df):
    """
    Apply professional Excel formatting:
    - Header styling (bold, background color)
    - Color-coded matching percentages
    - Auto-adjust column widths
    - Borders and alignment
    """
    
    # Define colors
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # 80-100%
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # 60-79%
    orange_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # 40-59%
    red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")     # <40%
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Format header row
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    
    # Format data rows
    for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row), start=2):
        for col_idx, cell in enumerate(row, start=1):
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            
            # Column B is "Matching Percentage" - color code it
            if col_idx == 2:  # Matching Percentage column
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(bold=True, size=11)
                
                try:
                    score = float(cell.value)
                    
                    # Color code based on score
                    if score >= 80:
                        cell.fill = green_fill
                    elif score >= 60:
                        cell.fill = yellow_fill
                    elif score >= 40:
                        cell.fill = orange_fill
                    else:
                        cell.fill = red_fill
                    
                    # Format as percentage
                    cell.number_format = '0.00"%"'
                
                except (ValueError, TypeError):
                    pass
    
    # Auto-adjust column widths
    auto_adjust_column_width(worksheet)
    
    # Freeze header row
    worksheet.freeze_panes = "A2"

def auto_adjust_column_width(worksheet):
    """
    Automatically adjust column widths based on content
    """
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    max_length = max(max_length, cell_length)
            except:
                pass
        
        # Set width with some padding (max 50 to avoid very wide columns)
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width

# ===============================
# SUMMARY STATISTICS SHEET
# ===============================
def create_summary_sheet(writer, df, job_role):
    """
    Create a summary sheet with statistics and insights
    """
    summary_data = {
        "Metric": [
            "Job Role",
            "Total Candidates",
            "Shortlisted (≥60%)",
            "Highly Qualified (≥80%)",
            "Average Score",
            "Highest Score",
            "Lowest Score",
            "Export Date",
            "Export Time"
        ],
        "Value": [
            job_role,
            len(df),
            len(df[df["Matching Percentage"] >= 60]),
            len(df[df["Matching Percentage"] >= 80]),
            f"{df['Matching Percentage'].mean():.2f}%",
            f"{df['Matching Percentage'].max():.2f}%",
            f"{df['Matching Percentage'].min():.2f}%",
            datetime.now().strftime("%Y-%m-%d"),
            datetime.now().strftime("%H:%M:%S")
        ]
    }
    
    summary_df = pd.DataFrame(summary_data)
    summary_df.to_excel(writer, sheet_name="Summary", index=False)
    
    # Format summary sheet
    workbook = writer.book
    summary_sheet = workbook["Summary"]
    
    # Apply formatting
    for cell in summary_sheet[1]:
        cell.font = Font(bold=True, size=12)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF", size=12)
    
    # Auto-adjust columns
    auto_adjust_column_width(summary_sheet)


# ===============================
# BATCH EXPORT (ALL FORMATS)
# ===============================
def export_all_formats(df, job_role="Resume Screening"):
    """
    Export in multiple formats at once
    Returns: dict with all export buffers
    """
    exports = {}
    
    # Excel
    exports['excel'] = export_excel(df, job_role)
    
    return exports
